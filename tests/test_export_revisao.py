# -*- coding: utf-8 -*-
"""Regressão: a exportação Excel deve respeitar as decisões da revisão (REV_STATE),
não reclassificar os itens crus. Cobre o bug em que /api/exportar_xlsx recebia
ITENS sem os flags _excluir/_motivo e contava itens intragrupo.
Roda como script: python tests/test_export_revisao.py
"""
import sys, os, io, shutil, tempfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pathlib import Path
import openpyxl
import app

falhas = []
def check(nome, cond, detalhe=""):
    print(("  OK  " if cond else "FALHA ") + nome + (f" — {detalhe}" if detalhe and not cond else ""))
    if not cond:
        falhas.append(nome)

_tmp = Path(tempfile.mkdtemp())
shutil.copy2(app.CFG_PATH, _tmp / "config.json")
app.CFG_PATH = _tmp / "config.json"
app.BACKUP_DIR = _tmp / "backup"
c = app.app.test_client()

# Item normal (entra) + item intragrupo marcado para exclusão (não entra)
ITENS = [
    {"desc": "TOMAHAWK", "cod": "", "peso": 1.0, "preco": 250.0, "total": 250.0,
     "cliente": "RESTAURANTE EXTERNO LTDA"},
    {"desc": "PICANHA", "cod": "", "peso": 2.0, "preco": 300.0, "total": 600.0,
     "cliente": "BEEF PASSION IND COM. DE ALIMENTOS LTDA EPP",
     "_excluir": True, "_motivo": "intragrupo"},
]
payload = {"itens": ITENS, "canal": "PJ", "nivel": 3, "mes": 5, "ano": 2026, "vendedor": "Teste"}

# 1. /api/calcular não conta o item intragrupo no faturamento COMPUTÁVEL
#    (fat_comp = base da comissão; fat_total = bruto, inclui tudo)
res = c.post("/api/calcular", json=payload).get_json()
check("calcular: item intragrupo fora do fat_comp (base da comissão)",
      abs(res["fat_comp"] - 250.0) < 1e-6, f'fat_comp={res["fat_comp"]}')

# 2. /api/exportar_xlsx respeita o _excluir (o xlsx marca Usado=Não / intragrupo)
r = c.post("/api/exportar_xlsx", json=payload)
check("exportar_xlsx → 200", r.status_code == 200)
wb = openpyxl.load_workbook(io.BytesIO(r.data), data_only=True)
ws = wb["BD_VENDAS_CLASSIFICADAS"]
# Cabeçalho na linha 4: ... Classificação(11) ... Faixa(16) Usado(17) Motivo(18)
linhas = {}
for rr in range(5, ws.max_row + 1):
    desc = str(ws.cell(rr, 7).value or "")
    if desc:
        linhas[desc.upper()] = {"faixa": ws.cell(rr, 16).value,
                                 "usado": ws.cell(rr, 17).value,
                                 "motivo": ws.cell(rr, 18).value}
pic = next((v for k, v in linhas.items() if "PICANHA" in k), None)
tom = next((v for k, v in linhas.items() if "TOMAHAWK" in k), None)
check("export: PICANHA (intragrupo) presente", pic is not None)
check("export: PICANHA marcada como NÃO usada",
      pic and str(pic["usado"]).upper().startswith("N"), str(pic))
check("export: PICANHA faixa/motivo intragrupo",
      pic and ("intragrupo" in str(pic["faixa"]).lower() or "intragrupo" in str(pic["motivo"]).lower()),
      str(pic))
check("export: TOMAHAWK (externo) usado no cálculo",
      tom and str(tom["usado"]).upper().startswith("S"), str(tom))

shutil.rmtree(_tmp, ignore_errors=True)
print()
if falhas:
    print(f"RESULTADO: {len(falhas)} FALHA(S) -> {falhas}")
    sys.exit(1)
print("RESULTADO: TODOS OS TESTES PASSARAM")
