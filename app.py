"""
BeefPassion · Calculadora OTE
Backend Flask + pdfplumber
Stoic Capital · v1.0
"""

import os
import json
import re
import tempfile
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file
import pdfplumber
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# ── Arquivo de configuração local ──────────────────────────────────────────────
CFG_PATH = Path(__file__).parent / "config.json"

PRECOS_TABELA_2026 = {
        "ACEM PESCOCO PECA|CLASSICO": 35.0,
        "ACEM PESCOCO PECA|RESERVA": 55.0,
        "ALCATRA PECA|CLASSICO": 89.0,
        "ALCATRA PECA|RESERVA": 95.0,
        "ANCHO STEAK|CLASSICO": 189.0,
        "ANCHO STEAK|RESERVA": 380.0,
        "ANCHO STEAK|SUPREME": 930.0,
        "ARANHA DA ALCATRA|RESERVA": 75.0,
        "ASSADO DE TIRA DIANTEIRO|CLASSICO": 161.0,
        "ASSADO DE TIRA DIANTEIRO|RESERVA": 218.0,
        "ASSADO DE TIRA TRASEIRO|CLASSICO": 110.0,
        "ASSADO DE TIRA TRASEIRO|RESERVA": 145.0,
        "BABY BEEF STEAK|CLASSICO": 89.0,
        "BABY BEEF STEAK|RESERVA": 142.0,
        "BANANINHA|CLASSICO": 132.0,
        "BANANINHA|RESERVA": 150.0,
        "BEEF STEAK|CLASSICO": 89.0,
        "BEEF STEAK|RESERVA": 125.0,
        "BEEF TALLOW 1.5KG|RESERVA": 170.0,
        "BEEF TALLOW 250G|RESERVA": 30.0,
        "BIFE DO VAZIO|CLASSICO": 120.0,
        "BIFE DO VAZIO|RESERVA": 172.0,
        "BIFITOS|CLASSICO": 69.0,
        "BIFITOS|RESERVA": 80.0,
        "BISTECA FIORENTINA|CLASSICO": 280.0,
        "BISTECA FIORENTINA|RESERVA": 437.0,
        "BISTECA SHORT RIB|CLASSICO": 194.0,
        "BISTECA SHORT RIB|RESERVA": 253.0,
        "BISTECA SHORT RIB|SUPREME": 830.0,
        "BOCHECHA|CLASSICO": 69.0,
        "BOCHECHA|RESERVA": 80.0,
        "BOICON BOVINO FATIADO|RESERVA": 80.0,
        "BOMBOM DA ALCATRA|CLASSICO": 131.0,
        "BOMBOM DA ALCATRA|RESERVA": 207.0,
        "BUCHO|CLASSICO": 45.0,
        "BUCHO|RESERVA": 69.0,
        "CALDO DE OSSO 100ML|RESERVA": 20.0,
        "CAPA DA PALETA|CLASSICO": 55.0,
        "CAPA DA PALETA|RESERVA": 70.0,
        "CAPA DE FILE STEAK|CLASSICO": 125.0,
        "CAPA DE FILE STEAK|RESERVA": 155.0,
        "CAPA DE FILE|CLASSICO": 115.0,
        "CAPA DE FILE|RESERVA": 140.0,
        "CECINA BRESAOLA|RESERVA": 60.0,
        "CHORIZO|CLASSICO": 169.0,
        "CHORIZO|RESERVA": 380.0,
        "CHORIZO|SUPREME": 930.0,
        "CONTRAFILE|CLASSICO": 160.0,
        "CONTRAFILE|RESERVA": 350.0,
        "CORACAO|RESERVA": 55.0,
        "COSTELA DE JANELA|CLASSICO": 145.0,
        "COSTELA DE JANELA|RESERVA": 182.0,
        "COSTELA DO DIANTEIRO C/OSSO|CLASSICO": 135.0,
        "COSTELA DO DIANTEIRO|CLASSICO": 120.0,
        "COSTELA DO DIANTEIRO|RESERVA": 178.0,
        "COSTELA MINGA|CLASSICO": 30.0,
        "COSTELA MINGA|RESERVA": 45.0,
        "COXAO DURO PECA|CLASSICO": 56.0,
        "COXAO DURO PECA|RESERVA": 110.0,
        "COXAO MOLE PECA|CLASSICO": 65.0,
        "COXAO MOLE PECA|RESERVA": 80.0,
        "CUBITOS|CLASSICO": 59.0,
        "CUBITOS|RESERVA": 79.0,
        "DENVER PASSION|CLASSICO": 390.0,
        "DENVER PASSION|RESERVA": 530.0,
        "DENVER PASSION|SUPREME": 930.0,
        "ENTRANHA PASSION|CLASSICO": 140.0,
        "ENTRANHA PASSION|RESERVA": 166.0,
        "FIGADO|RESERVA": 79.0,
        "FILE DE COSTELA|CLASSICO": 179.0,
        "FILE DE COSTELA|RESERVA": 380.0,
        "FILE MIGNON MEDALHAO|CLASSICO": 253.0,
        "FILE MIGNON MEDALHAO|RESERVA": 360.0,
        "FILE MIGNON PECA|CLASSICO": 241.0,
        "FILE MIGNON PECA|RESERVA": 322.0,
        "FLAT IRON PASSION|CLASSICO": 224.0,
        "FLAT IRON PASSION|RESERVA": 264.0,
        "FLAT IRON PASSION|SUPREME": 830.0,
        "FRALDA INTERNA|CLASSICO": 89.0,
        "FRALDA INTERNA|RESERVA": 159.0,
        "FRALDINHA PASSION STEAK|CLASSICO": 218.0,
        "FRALDINHA PASSION STEAK|RESERVA": 593.0,
        "FRALDINHA PASSION|CLASSICO": 212.0,
        "FRALDINHA PASSION|RESERVA": 286.0,
        "GORDURA PECA|RESERVA": 20.0,
        "HAMBURGUER 180G|RESERVA": 49.0,
        "HAMBURGUER DE COSTELA 180G|RESERVA": 40.0,
        "HAMBURGUER SMASH 50G|RESERVA": 45.0,
        "ISCAS MIGNON|CLASSICO": 115.0,
        "ISCAS MIGNON|RESERVA": 132.0,
        "ISCAS PASSION|CLASSICO": 60.0,
        "ISCAS PASSION|RESERVA": 74.0,
        "KNUCKLE PASSION STEAK|CLASSICO": 69.0,
        "KNUCKLE PASSION STEAK|RESERVA": 113.0,
        "LAGARTO|CLASSICO": 60.0,
        "LAGARTO|RESERVA": 89.0,
        "LINGUA|CLASSICO": 60.0,
        "LINGUA|RESERVA": 90.0,
        "LINGUICA BOVINA|RESERVA": 50.0,
        "MAMINHA|CLASSICO": 170.0,
        "MAMINHA|RESERVA": 253.0,
        "MAMINHA|SUPREME": 700.0,
        "MOIDA BLEND|CLASSICO": 40.0,
        "MOIDA BLEND|RESERVA": 60.0,
        "MOIDA PATINHO|CLASSICO": 65.0,
        "MOIDA PATINHO|RESERVA": 74.0,
        "MUSCULO EM CUBOS|CLASSICO": 35.0,
        "MUSCULO EM CUBOS|RESERVA": 55.0,
        "MUSCULO PECA|CLASSICO": 30.0,
        "MUSCULO PECA|RESERVA": 44.0,
        "OSSO MOCOTO|RESERVA": 18.0,
        "OSSOBUCO|CLASSICO": 42.0,
        "OSSOBUCO|RESERVA": 59.0,
        "PALETA PECA|CLASSICO": 89.0,
        "PALETA PECA|RESERVA": 128.0,
        "PASSIONITOS|CLASSICO": 69.0,
        "PASSIONITOS|RESERVA": 80.0,
        "PATINHO PECA|CLASSICO": 57.0,
        "PATINHO PECA|RESERVA": 99.0,
        "PEITO BRISKET|CLASSICO": 59.0,
        "PEITO BRISKET|RESERVA": 89.0,
        "PEIXINHO|CLASSICO": 59.0,
        "PEIXINHO|RESERVA": 70.0,
        "PICANHA PASSION|CLASSICO": 289.0,
        "PICANHA PASSION|RESERVA": 499.0,
        "PICANHA PASSION|SUPREME": 930.0,
        "RABADA|CLASSICO": 55.0,
        "RABADA|RESERVA": 79.0,
        "RECORTE|CLASSICO": 35.0,
        "RECORTE|RESERVA": 45.0,
        "SALSICHA BOVINA|RESERVA": 60.0,
        "SHOULDER STEAK|CLASSICO": 99.0,
        "SHOULDER STEAK|RESERVA": 135.0,
        "STEAK DA CASA|CLASSICO": 75.0,
        "STEAK DA CASA|RESERVA": 95.0,
        "STEAK DO ACEM COWBOY|CLASSICO": 99.0,
        "STEAK DO ACEM COWBOY|RESERVA": 149.0,
        "STEAK PASSION|CLASSICO": 65.0,
        "STEAK PASSION|RESERVA": 119.0,
        "STEAK PASSION|SUPREME": 320.0,
        "TBONE|CLASSICO": 210.0,
        "TBONE|RESERVA": 437.0,
        "TOMAHAWK|CLASSICO": 210.0,
        "TOMAHAWK|RESERVA": 417.0
}

# ── Configuração padrão ────────────────────────────────────────────────────────
DEFAULT_CFG = {
    "precos_pj": dict(PRECOS_TABELA_2026),  # Tabela 2026 PJ
    "precos_pf": dict(PRECOS_TABELA_2026),  # Tabela 2026 PF (inicialmente igual à PJ)
    "vendedores": [
        {"nome": "Bruno",    "canal": "PJ",    "cargo": "Comercial PJ",             "nivel_padrao": 3},
        {"nome": "Anderson", "canal": "PF",    "cargo": "Gerente / Vendedor PF",    "nivel_padrao": 3},
        {"nome": "Geovanna", "canal": "ADMIN", "cargo": "Auxiliar Adm. e Comercial","nivel_padrao": 3},
    ],
    "ote": {
        "PJ": [
            {"n":1,"ote":4200,"fixo":2520,"var":1680,"pf":0.6,"pv":0.4,"meta":200000,"corte":140000},
            {"n":2,"ote":5125,"fixo":3075,"var":2050,"pf":0.6,"pv":0.4,"meta":250000,"corte":175000},
            {"n":3,"ote":6000,"fixo":3600,"var":2400,"pf":0.6,"pv":0.4,"meta":300000,"corte":210000},
            {"n":4,"ote":6825,"fixo":4095,"var":2730,"pf":0.6,"pv":0.4,"meta":350000,"corte":245000},
            {"n":5,"ote":8662.5,"fixo":5197.5,"var":3465,"pf":0.6,"pv":0.4,"meta":450000,"corte":315000},
            {"n":6,"ote":10450,"fixo":6270,"var":4180,"pf":0.6,"pv":0.4,"meta":550000,"corte":385000},
            {"n":7,"ote":12187.5,"fixo":7312.5,"var":4875,"pf":0.6,"pv":0.4,"meta":650000,"corte":455000},
            {"n":8,"ote":13875,"fixo":8325,"var":5550,"pf":0.6,"pv":0.4,"meta":750000,"corte":525000},
            {"n":9,"ote":16425,"fixo":9855,"var":6570,"pf":0.6,"pv":0.4,"meta":900000,"corte":630000},
            {"n":10,"ote":18900,"fixo":11340,"var":7560,"pf":0.6,"pv":0.4,"meta":1050000,"corte":735000},
        ],
        "PF": [
            {"n":1,"ote":3000,"fixo":2100,"var":900,"pf":0.7,"pv":0.3,"meta":75000,"corte":52500},
            {"n":2,"ote":3950,"fixo":2765,"var":1185,"pf":0.7,"pv":0.3,"meta":100000,"corte":70000},
            {"n":3,"ote":4875,"fixo":3412.5,"var":1462.5,"pf":0.7,"pv":0.3,"meta":125000,"corte":87500},
            {"n":4,"ote":5775,"fixo":4042.5,"var":1732.5,"pf":0.7,"pv":0.3,"meta":150000,"corte":105000},
            {"n":5,"ote":6650,"fixo":4655,"var":1995,"pf":0.7,"pv":0.3,"meta":175000,"corte":122500},
            {"n":6,"ote":7500,"fixo":5250,"var":2250,"pf":0.7,"pv":0.3,"meta":200000,"corte":140000},
            {"n":7,"ote":8325,"fixo":5827.5,"var":2497.5,"pf":0.7,"pv":0.3,"meta":225000,"corte":157500},
            {"n":8,"ote":9125,"fixo":6387.5,"var":2737.5,"pf":0.7,"pv":0.3,"meta":250000,"corte":175000},
            {"n":9,"ote":9900,"fixo":6930,"var":2970,"pf":0.7,"pv":0.3,"meta":275000,"corte":192500},
            {"n":10,"ote":10650,"fixo":7455,"var":3195,"pf":0.7,"pv":0.3,"meta":300000,"corte":210000},
        ],
        "ADMIN": [
            {"n":1,"ote":2337.5,"fixo":1636.25,"var":701.25,"pf":0.7,"pv":0.3,"meta":275000,"corte":192500},
            {"n":2,"ote":2870,"fixo":2009,"var":861,"pf":0.7,"pv":0.3,"meta":350000,"corte":245000},
            {"n":3,"ote":3357.5,"fixo":2350.25,"var":1007.25,"pf":0.7,"pv":0.3,"meta":425000,"corte":297500},
            {"n":4,"ote":3800,"fixo":2660,"var":1140,"pf":0.7,"pv":0.3,"meta":500000,"corte":350000},
            {"n":5,"ote":4562.5,"fixo":3193.75,"var":1368.75,"pf":0.7,"pv":0.3,"meta":625000,"corte":437500},
            {"n":6,"ote":5250,"fixo":3675,"var":1575,"pf":0.7,"pv":0.3,"meta":750000,"corte":525000},
            {"n":7,"ote":5862.5,"fixo":4103.75,"var":1758.75,"pf":0.7,"pv":0.3,"meta":875000,"corte":612500},
            {"n":8,"ote":6400,"fixo":4480,"var":1920,"pf":0.7,"pv":0.3,"meta":1000000,"corte":700000},
            {"n":9,"ote":7167.5,"fixo":5017.25,"var":2150.25,"pf":0.7,"pv":0.3,"meta":1175000,"corte":822500},
            {"n":10,"ote":7830,"fixo":5481,"var":2349,"pf":0.7,"pv":0.3,"meta":1350000,"corte":945000},
        ],
    },
    "senha": "beefpassion",
    "mult_table": {}  # sobrescritas do usuário — chave "1.00" → {c1,c2,c3,c4}
}

# Tabela de multiplicadores exata
# Fonte: Beef_Metas_OTE.xlsx · aba "1. Tabela de Multiplicador"
# Cols: [atingimento, Abaixo10%, 3-10%Desc, IdealLE3%, AcimaTabela]
MULT_TABLE = [
    [0,0,0,0,0],[0.01,0,0,0,0],[0.02,0,0,0,0],[0.03,0,0,0,0],[0.04,0,0,0,0],
    [0.05,0,0,0,0],[0.06,0,0,0,0],[0.07,0,0,0,0],[0.08,0,0,0,0],[0.09,0,0,0,0],
    [0.1,0,0,0,0],[0.11,0,0,0,0],[0.12,0,0,0,0],[0.13,0,0,0,0],[0.14,0,0,0,0],
    [0.15,0,0,0,0],[0.16,0,0,0,0],[0.17,0,0,0,0],[0.18,0,0,0,0],[0.19,0,0,0,0],
    [0.2,0,0,0,0],[0.21,0,0,0,0],[0.22,0,0,0,0],[0.23,0,0,0,0],[0.24,0,0,0,0],
    [0.25,0,0,0,0],[0.26,0,0,0,0],[0.27,0,0,0,0],[0.28,0,0,0,0],[0.29,0,0,0,0],
    [0.3,0,0,0,0],[0.31,0,0,0,0],[0.32,0,0,0,0],[0.33,0,0,0,0],[0.34,0,0,0,0],
    [0.35,0,0,0,0],[0.36,0,0,0,0],[0.37,0,0,0,0],[0.38,0,0,0,0],[0.39,0,0,0,0],
    [0.4,0,0,0,0],[0.41,0,0,0,0],[0.42,0,0,0,0],[0.43,0,0,0,0],[0.44,0,0,0,0],
    [0.45,0,0,0,0],[0.46,0,0,0,0],[0.47,0,0,0,0],[0.48,0,0,0,0],[0.49,0,0,0,0],
    [0.5,0,0,0,0],[0.51,0,0,0,0],[0.52,0,0,0,0],[0.53,0,0,0,0],[0.54,0,0,0,0],
    [0.55,0,0,0,0],[0.56,0,0,0,0],[0.57,0,0,0,0],[0.58,0,0,0,0],[0.59,0,0,0,0],
    [0.6,0,0,0,0],[0.61,0,0,0,0],[0.62,0,0,0,0],[0.63,0,0,0,0],[0.64,0,0,0,0],
    [0.65,0,0,0,0],[0.66,0,0,0,0],[0.67,0,0,0,0],[0.68,0,0,0,0],[0.69,0,0,0,0],
    [0.69,0,0,0,0],[0.7,-0.3,0.4,0.5,0.6],[0.71,-0.32,0.41,0.51,0.615],
    [0.72,-0.34,0.42,0.52,0.64],[0.73,-0.36,0.43,0.53,0.66],[0.74,-0.38,0.44,0.54,0.68],
    [0.75,-0.4,0.45,0.55,0.7],[0.76,-0.42,0.46,0.56,0.72],[0.77,-0.44,0.47,0.57,0.74],
    [0.78,-0.46,0.48,0.58,0.76],[0.79,-0.48,0.49,0.59,0.78],[0.8,-0.5,0.5,0.605,0.8],
    [0.81,-0.52,0.51,0.62,0.82],[0.82,-0.54,0.52,0.635,0.84],[0.83,-0.56,0.53,0.65,0.86],
    [0.84,-0.58,0.54,0.665,0.88],[0.85,-0.6,0.55,0.68,0.9],[0.86,-0.62,0.56,0.695,0.92],
    [0.87,-0.64,0.57,0.71,0.94],[0.88,-0.66,0.58,0.725,0.96],[0.89,-0.68,0.59,0.725,0.98],
    [0.9,-0.7,0.6,0.75,1],[0.91,-0.72,0.61,0.775,1.02],[0.92,-0.74,0.62,0.8,1.04],
    [0.93,-0.76,0.63,0.825,1.06],[0.94,-0.78,0.64,0.85,1.08],[0.95,-0.8,0.65,0.875,1.1],
    [0.96,-0.82,0.66,0.9,1.12],[0.97,-0.84,0.67,0.925,1.14],[0.98,-0.86,0.68,0.95,1.16],
    [0.99,-0.88,0.69,0.975,1.18],[1,-0.9,0.7,1,1.2],[1.01,-0.91,0.71,1.01,1.21],
    [1.02,-0.92,0.72,1.02,1.22],[1.03,-0.93,0.73,1.03,1.23],[1.04,-0.94,0.74,1.04,1.24],
    [1.05,-0.95,0.75,1.05,1.25],[1.06,-0.96,0.76,1.06,1.26],[1.07,-0.97,0.77,1.07,1.27],
    [1.08,-0.98,0.78,1.08,1.28],[1.09,-0.99,0.79,1.09,1.29],[1.1,-1,0.8,1.1,1.3],
    [1.11,-1.01,0.81,1.11,1.31],[1.12,-1.02,0.82,1.12,1.32],[1.13,-1.03,0.83,1.13,1.33],
    [1.14,-1.04,0.84,1.14,1.34],[1.15,-1.05,0.85,1.15,1.35],[1.16,-1.06,0.86,1.16,1.36],
    [1.17,-1.07,0.87,1.17,1.37],[1.18,-1.08,0.88,1.18,1.38],[1.19,-1.09,0.89,1.19,1.39],
    [1.2,-1.1,0.9,1.2,1.4],[1.21,-1.11,0.91,1.21,1.41],[1.22,-1.12,0.92,1.22,1.42],
    [1.23,-1.13,0.93,1.23,1.43],[1.24,-1.14,0.94,1.24,1.44],[1.25,-1.15,0.95,1.25,1.45],
    [1.26,-1.16,0.96,1.26,1.46],[1.27,-1.17,0.97,1.27,1.47],[1.28,-1.18,0.98,1.28,1.48],
    [1.29,-1.19,0.99,1.29,1.49],[1.3,-1.2,0.99,1.29,1.5],[1.31,-1.21,1.0,1.3,1.51],
    [1.32,-1.22,1.01,1.31,1.52],[1.33,-1.23,1.02,1.32,1.53],[1.34,-1.24,1.03,1.33,1.54],
    [1.35,-1.25,1.04,1.34,1.55],[1.36,-1.26,1.05,1.35,1.56],[1.37,-1.27,1.06,1.36,1.57],
    [1.38,-1.28,1.07,1.37,1.58],[1.39,-1.29,1.08,1.38,1.59],[1.4,-1.3,1.09,1.39,1.6],
    [1.5,-1.4,1.19,1.49,1.7],[1.6,-1.5,1.29,1.59,1.8],[1.7,-1.6,1.39,1.69,1.9],
    [1.8,-1.7,1.49,1.79,2.0],[1.9,-1.8,1.59,1.89,2.1],[2.0,-1.9,1.69,1.99,2.2],
    [2.1,-2.0,1.79,2.09,2.3],[2.2,-2.1,1.89,2.19,2.4],[2.3,-2.2,1.99,2.29,2.5],
    [2.4,-2.3,2.09,2.39,2.6],[2.5,-2.4,2.19,2.49,2.7],
]

def get_mult(ating: float, col: int, mult_override: dict = None) -> float:
    """Lookup na tabela de multiplicadores. col: 1=Abaixo10%, 2=3-10%Desc, 3=Ideal, 4=Acima.
    mult_override: dict chave '1.00' → {c1,c2,c3,c4} — sobrescritas do usuário.
    """
    if col < 1 or col > 4:
        return 0.0
    a = min(max(ating, 0.0), 2.5)
    # Busca binária pelo atingimento mais próximo
    best_row = MULT_TABLE[0]
    for row in MULT_TABLE:
        if row[0] <= a:
            best_row = row
        else:
            break
    # Verificar override
    if mult_override:
        key = f"{best_row[0]:.2f}"
        ov  = mult_override.get(key)
        if ov:
            return float(ov.get(f"c{col}", 0) or 0)
    return best_row[col] or 0.0


# ── Config helpers ─────────────────────────────────────────────────────────────
def load_cfg() -> dict:
    if CFG_PATH.exists():
        try:
            with open(CFG_PATH, encoding="utf-8") as f:
                saved = json.load(f)
            cfg = json.loads(json.dumps(DEFAULT_CFG))
            cfg.update(saved)
            # Backward compat: migrar "precos" legado → precos_pj + precos_pf
            if "precos" in cfg and "precos_pj" not in cfg:
                cfg["precos_pj"] = cfg.pop("precos")
                cfg["precos_pf"] = dict(cfg["precos_pj"])
            elif "precos" in cfg:
                cfg.pop("precos", None)  # remove legado se já tem pj/pf
            return cfg
        except Exception:
            pass
    return json.loads(json.dumps(DEFAULT_CFG))


def save_cfg(cfg: dict):
    with open(CFG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ── PDF Extraction ─────────────────────────────────────────────────────────────
def extrair_itens(pdf_path: str) -> dict:
    """
    Extrai itens de venda do relatório de faturamento BeefPassion (FastReport).
    Usa pdfplumber com extração por coordenadas para lidar com layout posicional.
    Retorna dict com itens extraídos e log de diagnóstico.
    """
    itens = []
    avisos = []
    paginas_ok = 0
    paginas_sem_item = 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pags = len(pdf.pages)

        for num_pag, page in enumerate(pdf.pages, 1):
            try:
                # Extração com tabelas (preferencial para layouts tabulares)
                tables = page.extract_tables({
                    "vertical_strategy": "lines_strict",
                    "horizontal_strategy": "lines_strict",
                    "snap_tolerance": 4,
                    "join_tolerance": 3,
                    "edge_min_length": 10,
                    "min_words_vertical": 1,
                })

                itens_pag = _parse_tables(tables, num_pag)

                # Fallback: extração por texto se tabelas não produziram resultado
                if not itens_pag:
                    words = page.extract_words(
                        x_tolerance=3, y_tolerance=3,
                        keep_blank_chars=False, use_text_flow=False
                    )
                    itens_pag = _parse_words(words, num_pag)

                if itens_pag:
                    itens.extend(itens_pag)
                    paginas_ok += 1
                else:
                    paginas_sem_item += 1

            except Exception as e:
                avisos.append(f"Página {num_pag}: erro de extração — {str(e)}")

    # Deduplicação por (pedido + cod + peso)
    vistos = set()
    itens_dedup = []
    for it in itens:
        chave = (it.get("pedido",""), it.get("cod",""), str(round(it.get("peso",0),2)))
        if chave not in vistos:
            vistos.add(chave)
            itens_dedup.append(it)

    return {
        "itens": itens_dedup,
        "total_paginas": total_pags,
        "paginas_com_itens": paginas_ok,
        "paginas_sem_item": paginas_sem_item,
        "avisos": avisos,
        "total_itens": len(itens_dedup),
    }


def _parse_valor_br(s: str) -> float:
    """Converte string BR (1.234,56) para float."""
    try:
        s = s.strip().replace("R$", "").strip()
        s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _parse_tables(tables, num_pag: int) -> list:
    """Tenta extrair itens de estrutura de tabela detectada pelo pdfplumber."""
    itens = []
    if not tables:
        return itens

    # Padrões de código de produto: 5 dígitos
    cod_re = re.compile(r"^\d{5}$")
    valor_re = re.compile(r"^[\d]{1,3}(?:\.\d{3})*,\d{2}$")

    for table in tables:
        for row in table:
            if not row:
                continue
            cells = [str(c).strip() if c else "" for c in row]
            # Identificar linha de item: primeira célula é código 5 dígitos
            if len(cells) >= 5 and cod_re.match(cells[0]):
                try:
                    cod = cells[0]
                    desc = cells[1] if len(cells) > 1 else ""
                    # Buscar campos numéricos nas últimas colunas
                    nums = []
                    for c in cells[2:]:
                        if valor_re.match(c) or re.match(r"^\d+,\d+$", c):
                            nums.append(_parse_valor_br(c))
                    if len(nums) >= 2:
                        peso = nums[0]
                        preco = nums[1] if len(nums) >= 3 else 0.0
                        total = nums[-1]
                        if total > 0:
                            itens.append({
                                "pag": num_pag,
                                "data": "", "pedido": "", "nf": "", "cliente": "",
                                "cod": cod, "desc": desc,
                                "peso": peso, "preco": preco, "total": total,
                            })
                except Exception:
                    continue
    return itens


def _parse_words(words: list, num_pag: int) -> list:
    """
    Extração por posição de palavras (fallback).
    Agrupa palavras por linha (Y) e tenta identificar padrões de item.
    """
    itens = []
    if not words:
        return itens

    # Agrupar por linha (Y top, tolerância 3pt)
    linhas = {}
    for w in words:
        y = round(w["top"] / 3) * 3
        linhas.setdefault(y, []).append(w)

    cod_re = re.compile(r"^\d{5}$")
    val_re = re.compile(r"^[\d\.]+,\d{2}$")

    ctx = {"data": "", "pedido": "", "nf": "", "cliente": ""}

    for y in sorted(linhas.keys()):
        tokens = sorted(linhas[y], key=lambda w: w["x0"])
        txts = [t["text"].strip() for t in tokens]
        linha = " ".join(txts)

        # Capturar data
        dm = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", linha)
        if dm:
            ctx["data"] = dm.group(1)

        # Capturar pedido + cliente
        pm = re.search(r"Pedido[:\s]+([\w\-]+)\s+Cliente[:\s]+\d+\s*[-–]\s*(.+?)(?:Nota|NF|$)", linha)
        if pm:
            ctx["pedido"] = pm.group(1).strip()
            ctx["cliente"] = pm.group(2).strip()[:60]

        # Capturar NF
        nfm = re.search(r"Nota\s*fiscal[:\s]+(\d+)", linha, re.IGNORECASE)
        if nfm:
            ctx["nf"] = nfm.group(1)

        # Linha de item de produto
        if txts and cod_re.match(txts[0]):
            nums = [_parse_valor_br(t) for t in txts if val_re.match(t)]
            desc_parts = [t for t in txts[1:] if not re.match(r"^[\d\.,R$%]+$", t)]
            desc = " ".join(desc_parts)[:60]
            if len(nums) >= 2 and nums[-1] > 0:
                peso  = nums[0]
                preco = nums[1] if len(nums) >= 3 else 0.0
                total = nums[-1]
                itens.append({
                    "pag": num_pag,
                    "data":    ctx["data"],
                    "pedido":  ctx["pedido"],
                    "nf":      ctx["nf"],
                    "cliente": ctx["cliente"],
                    "cod":     txts[0],
                    "desc":    desc,
                    "peso":    peso,
                    "preco":   preco,
                    "total":   total,
                })
    return itens


# ── Cálculo OTE ────────────────────────────────────────────────────────────────
FAIXA_COLS = {"abaixo": 1, "desconto": 2, "ideal": 3, "acima": 4}
FAIXA_LABELS = {
    "abaixo":  "Abaixo 10%",
    "desconto":"3 a 10% Desc.",
    "ideal":   "Tabela Ideal (≤3%)",
    "acima":   "Acima da Tabela",
    "sem_ref": "Sem Referência",
}

# ── Stopwords para matching de descrições do PDF ────────────────────────────────
# Palavras que aparecem nas descrições do ERP mas NÃO nas chaves da tabela de preços
_MATCH_STOP = {
    "CONG","CG","WAGYU","PASSION","B","S","SOB","ENCOMENDA",
    "FATIADO","PICADO","CLASSICO","/","(",")","-","E","DA","DE","DO","DAS","DOS","EM","A"
}

# Índice pré-computado por canal: {"PJ": {cat→[...]}, "PF": {cat→[...]}}
# Construído na primeira chamada de cada canal e cacheado
_MATCH_INDEX: dict = {}

def _tokens(s: str) -> set:
    """Tokeniza string para matching: uppercase, remove pontuação e stopwords."""
    import re
    s = re.sub(r"[/()\-]", " ", s.upper())
    return {w for w in s.split() if w not in _MATCH_STOP and len(w) > 1}

def _build_match_index(precos: dict) -> dict:
    idx: dict = {}
    for chave, preco in precos.items():
        if "|" not in chave:
            continue
        desc_tab, cat = chave.rsplit("|", 1)
        tks = _tokens(desc_tab)
        idx.setdefault(cat, []).append((tks, desc_tab, chave, preco))
    return idx

def _inferir_categoria(desc: str) -> str:
    d = desc.upper()
    if "SUPREME" in d:
        return "SUPREME"
    if any(t in d for t in ("RESERVA","RES ","BP 3","BP 4","BP 5","BP 6","BP3","BP4","BP5","BP6"," RES")):
        return "RESERVA"
    return "CLASSICO"

def _match_preco(desc_pdf: str, precos: dict, canal: str = "PJ"):
    """
    Encontra o melhor match na tabela de preços para uma descrição do PDF.
    Retorna (preco_ref, chave_match, categoria, score) ou (None, None, None, 0).
    Algoritmo: score = 0.7 * cobertura_tabela + 0.3 * cobertura_pdf, threshold=0.45
    """
    global _MATCH_INDEX
    if canal not in _MATCH_INDEX:
        _MATCH_INDEX[canal] = _build_match_index(precos)

    cat_inf = _inferir_categoria(desc_pdf)
    tks_pdf = _tokens(desc_pdf)
    if not tks_pdf:
        return None, None, None, 0

    idx = _MATCH_INDEX[canal]
    melhor = (None, None, None, 0)
    for cat_try in [cat_inf] + [c for c in ("CLASSICO","RESERVA","SUPREME") if c != cat_inf]:
        for tks_tab, desc_tab, chave, preco in idx.get(cat_try, []):
            inter = len(tks_pdf & tks_tab)
            if inter == 0:
                continue
            cob_tab = inter / max(len(tks_tab), 1)
            cob_pdf = inter / max(len(tks_pdf), 1)
            score   = cob_tab * 0.7 + cob_pdf * 0.3
            if score > melhor[3]:
                melhor = (preco, chave, cat_try, score)
        # Se já encontrou bom match na categoria inferida, não forçar fallback
        if melhor[3] >= 0.70 and cat_try == cat_inf:
            break

    if melhor[3] >= 0.45:
        return melhor[0], melhor[1], melhor[2], melhor[3]
    return None, None, None, 0


def classificar_item(item: dict, precos: dict, canal: str = "PJ") -> dict:
    """
    Classifica um item por faixa de preço.
    Respeita overrides vindos do frontend (revisão manual):
      _excluir          → faixa "excluido" (não entra no cálculo)
      _motivo           → motivo da exclusão: 'fora_competencia' | 'excluido_operador'
      _preco_ref_override → usa este preço como referência em vez do match automático
      _match_key        → chave explícita da tabela escolhida pelo operador
    """
    # ── Exclusões explícitas do operador / competência ───────────────────────
    if item.get("_excluir"):
        motivo = item.get("_motivo", "excluido_operador")
        faixa  = "fora_comp" if motivo == "fora_competencia" else "excluido"
        return {**item, "faixa": faixa, "desvio": None, "preco_ref": None,
                "preco_key": None, "categoria": _inferir_categoria(item.get("desc","")),
                "match_score": 0.0, "_motivo": motivo}

    # ── Override manual de preço de referência (vincular / avulso) ───────────
    ref_override  = item.get("_preco_ref_override")
    match_override = item.get("_match_key")

    if ref_override and ref_override > 0:
        ref        = float(ref_override)
        chave_match = match_override or "AVULSO"
        cat_match   = _inferir_categoria(item.get("desc",""))
        score       = 1.0
    else:
        # Match automático
        desc_pdf = item.get("desc", "")
        ref, chave_match, cat_match, score = _match_preco(desc_pdf, precos, canal)

    if not ref or ref <= 0:
        return {**item, "faixa": "sem_ref", "desvio": None, "preco_ref": None,
                "preco_key": None, "categoria": _inferir_categoria(item.get("desc","")),
                "match_score": 0.0}

    desvio = (item["preco"] - ref) / ref
    if desvio > 0:
        faixa = "acima"
    elif desvio >= -0.03:
        faixa = "ideal"
    elif desvio >= -0.10:
        faixa = "desconto"
    else:
        faixa = "abaixo"

    return {**item, "faixa": faixa, "desvio": round(desvio, 6), "preco_ref": ref,
            "preco_key": chave_match, "categoria": cat_match, "match_score": round(score, 3)}


def calcular_ote(itens: list, ote_row: dict, precos: dict, canal: str = "PJ") -> dict:
    classified = [classificar_item(it, precos, canal) for it in itens]

    # Faixas que entram no cálculo do variável
    FAIXAS_COMP = ["abaixo", "desconto", "ideal", "acima"]
    # Faixas excluídas — aparecem no extrato mas não contam
    FAIXAS_EXCL = ["sem_ref", "excluido", "fora_comp"]

    fx = {f: {"fat": 0.0, "itens": []} for f in FAIXAS_COMP + FAIXAS_EXCL}
    for it in classified:
        faixa = it["faixa"]
        if faixa not in fx:
            fx[faixa] = {"fat": 0.0, "itens": []}
        fx[faixa]["fat"]   += it["total"]
        fx[faixa]["itens"].append(it)

    fat_excl  = sum(fx[f]["fat"] for f in FAIXAS_EXCL if f in fx)
    fat_total = sum(fx[f]["fat"] for f in fx)
    fat_comp  = fat_total - fat_excl
    ating     = fat_comp / ote_row["meta"] if ote_row["meta"] > 0 else 0.0

    r = {}
    mult_override = precos.get('__mult_override__')  # passado via contexto se disponível
    for nome in FAIXAS_COMP:
        fat = fx[nome]["fat"]
        if fat <= 0 or fat_comp <= 0:
            r[nome] = {"fat": fat, "pct": 0.0, "var_fx": 0.0, "mult": 0.0}
            continue
        col  = FAIXA_COLS[nome]
        mult = get_mult(ating, col, mult_override)
        prop = fat / fat_comp
        r[nome] = {"fat": fat, "pct": prop, "var_fx": prop * ote_row["var"] * mult, "mult": mult}

    # Grupo de excluídos consolidado para o relatório
    r["sem_ref"]  = {"fat": fx["sem_ref"]["fat"],  "pct": 0.0, "var_fx": 0.0, "mult": 0.0}
    r["excluido"] = {"fat": fx["excluido"]["fat"],  "pct": 0.0, "var_fx": 0.0, "mult": 0.0}
    r["fora_comp"]= {"fat": fx["fora_comp"]["fat"], "pct": 0.0, "var_fx": 0.0, "mult": 0.0}

    var_total  = sum(r[f]["var_fx"] for f in FAIXAS_COMP)
    var_final  = max(var_total, 0.0)
    rem_total  = ote_row["fixo"] + var_final

    # Contadores de auditoria
    n_comp  = sum(1 for it in classified if it["faixa"] in FAIXAS_COMP)
    n_excl  = sum(1 for it in classified if it["faixa"] in FAIXAS_EXCL)
    n_fora  = sum(1 for it in classified if it["faixa"] == "fora_comp")
    n_op    = sum(1 for it in classified if it["faixa"] == "excluido")
    n_semref= sum(1 for it in classified if it["faixa"] == "sem_ref")

    return {
        "classified":   classified,
        "fx":           fx,
        "r":            r,
        "fat_total":    fat_total,
        "fat_comp":     fat_comp,
        "ating":        ating,
        "var_total":    var_total,
        "var_final":    var_final,
        "fixo":         ote_row["fixo"],
        "rem_total":    rem_total,
        "atinge_meta":  fat_comp >= ote_row["meta"],
        "atinge_corte": fat_comp >= ote_row["corte"],
        "auditoria": {
            "total":        len(classified),
            "computados":   n_comp,
            "excluidos":    n_excl,
            "fora_comp":    n_fora,
            "excl_operador":n_op,
            "sem_ref":      n_semref,
        }
    }


# ── Excel Export ───────────────────────────────────────────────────────────────
NAVY   = "0D1838"
GOLD   = "DCC697"
CREAM  = "F7F4EE"
GREEN  = "1A6B3C"
RED    = "C0392B"
ORANGE = "D4620A"
BLUE3  = "1F4E79"
SMOKE  = "F0EDE6"
WHITE  = "FFFFFF"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

FMT_BRL  = 'R$ #,##0.00'
FMT_PCT  = '0.0%'
FMT_NUM  = '#,##0.000'


def gerar_xlsx(resultado: dict, vendedor: dict, nivel: int, mes: int, ano: int) -> str:
    """Gera arquivo Excel e retorna o caminho temporário."""
    wb = openpyxl.Workbook()

    # ── Aba 1: Extrato ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "CP_EXTRATO"
    ws.sheet_view.showGridLines = False

    MESES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

    ote_row_ref = resultado["ote_row"]
    r   = resultado["r"]
    fat = resultado["fat_comp"]
    ating = resultado["ating"]

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = f"EXTRATO DE COMISSÃO OTE — {vendedor['nome']} — {MESES[mes]}/{ano} — Nível {nivel}"
    ws["A1"].fill   = _fill(NAVY)
    ws["A1"].font   = _font(bold=True, color=GOLD, size=13)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{vendedor['cargo']} · Modelo OTE · Fonte: _Beef__Metas_OTE.pdf · Stoic Capital · 26/12/2025"
    ws["A2"].font = _font(color="888888", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Resumo financeiro
    row = 4
    headers_resumo = ["Item", "Valor"]
    for col, h in enumerate(headers_resumo, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _fill(NAVY); c.font = _font(bold=True, color=GOLD, size=10)
        c.alignment = _center(); c.border = _border_thin()
    ws.row_dimensions[row].height = 22

    resumo_dados = [
        ("Faturamento Computável",      fat,                          FMT_BRL),
        ("Meta Mensal (Nível %d)" % nivel, ote_row_ref["meta"],       FMT_BRL),
        ("Atingimento",                 ating,                        FMT_PCT),
        ("Corte (70%)",                 ote_row_ref["corte"],          FMT_BRL),
        ("Salário Fixo",                resultado["fixo"],             FMT_BRL),
        ("Variável a Receber",          resultado["var_final"],        FMT_BRL),
        ("REMUNERAÇÃO TOTAL",           resultado["rem_total"],        FMT_BRL),
    ]
    for i, (label, val, fmt) in enumerate(resumo_dados):
        r_idx = row + 1 + i
        is_total = (i == len(resumo_dados) - 1)
        lc = ws.cell(row=r_idx, column=1, value=label)
        vc = ws.cell(row=r_idx, column=2, value=val)
        vc.number_format = fmt
        lc.border = vc.border = _border_thin()
        if is_total:
            lc.fill = vc.fill = _fill(NAVY)
            lc.font = _font(bold=True, color=GOLD, size=11)
            vc.font = _font(bold=True, color=GOLD, size=11)
        else:
            bg = SMOKE if i % 2 == 0 else WHITE
            lc.fill = vc.fill = _fill(bg)
            lc.font = _font(size=10)
            vc.font = _font(size=10)
        vc.alignment = _right()
        ws.row_dimensions[r_idx].height = 18

    # Desdobramento por faixa
    row_fx = row + len(resumo_dados) + 3
    ws.merge_cells(f"A{row_fx}:H{row_fx}")
    ws.cell(row=row_fx, column=1, value="DESDOBRAMENTO POR FAIXA DE PREÇO")
    ws.cell(row=row_fx, column=1).fill = _fill(BLUE3)
    ws.cell(row=row_fx, column=1).font = _font(bold=True, color=WHITE, size=10)
    ws.cell(row=row_fx, column=1).alignment = _center()
    ws.row_dimensions[row_fx].height = 22

    row_fx += 1
    fx_headers = ["Faixa", "Faturamento (R$)", "% Computável", "Multiplicador", "Variável (R$)"]
    fx_colors  = [NAVY, RED, ORANGE, BLUE3, GREEN, "888888"]
    for col, h in enumerate(fx_headers, 1):
        c = ws.cell(row=row_fx, column=col, value=h)
        c.fill = _fill(NAVY); c.font = _font(bold=True, color=GOLD, size=10)
        c.alignment = _center(); c.border = _border_thin()

    FAIXA_DISPLAY = [
        ("abaixo",   "⬇ Abaixo 10%",       RED),
        ("desconto", "↘ 3 a 10% Desc.",     ORANGE),
        ("ideal",    "✓ Tabela Ideal (≤3%)",BLUE3),
        ("acima",    "⬆ Acima da Tabela",   GREEN),
        ("sem_ref",  "— Sem Referência",    "888888"),
    ]
    for i, (faixa, label, cor) in enumerate(FAIXA_DISPLAY):
        rr = row_fx + 1 + i
        rd = resultado["r"][faixa]
        vals = [label, rd["fat"], rd["pct"], rd.get("mult", 0.0), rd["var_fx"]]
        fmts = [None, FMT_BRL, FMT_PCT, "0.0000", FMT_BRL]
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=rr, column=col, value=val)
            if fmt: c.number_format = fmt
            c.border = _border_thin()
            bg = SMOKE if i % 2 == 0 else WHITE
            c.fill = _fill(bg)
            if col == 1:
                c.font = _font(bold=True, color=cor, size=10)
            else:
                c.font = _font(size=10)
                c.alignment = _right()
        ws.row_dimensions[rr].height = 18

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # ── Aba 2: Vendas Classificadas ────────────────────────────────────────────
    ws2 = wb.create_sheet("BD_VENDAS_CLASSIFICADAS")
    ws2.sheet_view.showGridLines = False

    # ── Linha 1: título ─────────────────────────────────────────────────────
    n_cols = 17  # total de colunas — atualizar se cols mudar
    ws2.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws2["A1"] = f"BD VENDAS CLASSIFICADAS — {vendedor['nome']} — {MESES[mes]}/{ano} — Nível {nivel}"
    ws2["A1"].fill = _fill(NAVY); ws2["A1"].font = _font(bold=True, color=GOLD, size=12)
    ws2["A1"].alignment = _center(); ws2.row_dimensions[1].height = 28

    # ── Linha 2: fonte e rastreabilidade ────────────────────────────────────
    from datetime import datetime as _dt
    ws2.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    ws2["A2"] = (
        f"Fonte: PDF de Faturamento BeefPassion (FastReport) | "
        f"Tabela de Preços: 00__2026_Tabela_PJ_Beef_Passion.pdf | "
        f"Período de Competência: {MESES[mes]}/{ano} | "
        f"Gerado em: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws2["A2"].font = _font(color="888888", size=9, italic=True)
    ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[2].height = 16

    # ── Linha 3: legenda de cores ────────────────────────────────────────────
    ws2.merge_cells(f"A3:{get_column_letter(n_cols)}3")
    ws2["A3"] = (
        "Legenda de cores → Verde: Acima da Tabela | "
        "Azul: Tabela Ideal (≤3% desc.) | "
        "Laranja: 3–10% Desconto | "
        "Vermelho: Abaixo 10% | "
        "Cinza: Sem Referência (NÃO entra no cálculo)"
    )
    ws2["A3"].font = _font(color="555555", size=8, italic=True)
    ws2["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[3].height = 14

    # ── Linha 4: cabeçalho das colunas ──────────────────────────────────────
    cols   = [
        "ID",
        "Data Competência",
        "Pedido",
        "NF",
        "Cliente",
        "Cód. Produto",
        "Descrição (PDF)",
        "Categoria",
        "Match Tabela (Chave)",
        "Peso (kg)",
        "Preço Praticado (R$/kg)",
        "Preço Ref. Tabela (R$/kg)",
        "Desvio %",
        "Faixa",
        "Usado no Cálculo",
        "Motivo Exclusão",
        "Total (R$)",
    ]
    widths = [8, 16, 12, 10, 32, 12, 42, 12, 36, 10, 22, 22, 10, 20, 18, 26, 14]

    for col, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.fill = _fill(NAVY); c.font = _font(bold=True, color=GOLD, size=10)
        c.alignment = _center(); c.border = _border_thin()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[4].height = 28
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:{get_column_letter(len(cols))}4"

    FAIXA_FILL = {
        "abaixo":   ("FFE8E8", RED),
        "desconto": ("FEF3E8", ORANGE),
        "ideal":    ("E8F0FE", BLUE3),
        "acima":    ("E8F5EE", GREEN),
        "sem_ref":  ("F0F0F0", "999999"),
        "excluido": ("F4F4F4", "AAAAAA"),
        "fora_comp":("EFEFEF", "AAAAAA"),
    }
    MOTIVO_LABEL = {
        "sem_ref":   "Sem referência de preço",
        "excluido":  "Excluído pelo operador",
        "fora_comp": "Fora da competência",
    }

    n_usado = 0
    n_excluido = 0

    for i, it in enumerate(resultado["classified"]):
        rr = 5 + i
        faixa  = it["faixa"]
        usado  = faixa not in ("sem_ref", "excluido", "fora_comp")
        bg, fc = FAIXA_FILL.get(faixa, (WHITE, "000000"))
        if usado: n_usado += 1
        else:     n_excluido += 1

        desvio_str = f"{it['desvio']*100:.1f}%" if it["desvio"] is not None else "—"
        usado_str  = "SIM" if usado else "NÃO"
        chave_match = it.get("preco_key") or "— sem match —"

        motivo_str = MOTIVO_LABEL.get(faixa, "") if not usado else ""
        row_data = [
            f"VND-{i+1:04d}",                        # A  ID
            it.get("data", ""),                       # B  Data Competência
            it.get("pedido", ""),                     # C  Pedido
            it.get("nf", ""),                         # D  NF
            it.get("cliente", ""),                    # E  Cliente
            it.get("cod", ""),                        # F  Cód. Produto
            it.get("desc", ""),                       # G  Descrição PDF
            it.get("categoria", "—"),                 # H  Categoria
            chave_match,                              # I  Match Tabela
            it.get("peso", 0.0),                      # J  Peso kg
            it.get("preco", 0.0),                     # K  Preço Praticado
            it.get("preco_ref") or "",                # L  Preço Ref. Tabela
            desvio_str,                               # M  Desvio %
            FAIXA_LABELS.get(faixa, faixa),           # N  Faixa
            usado_str,                                # O  Usado no Cálculo
            motivo_str,                               # P  Motivo Exclusão
            it.get("total", 0.0),                     # Q  Total R$
        ]
        fmts_row = [
            None, None, None, None, None, None, None, None, None,
            FMT_NUM, FMT_BRL, FMT_BRL, None, None, None, None, FMT_BRL
        ]
        aligns_right = {10, 11, 12, 16}  # colunas com alinhamento à direita (1-based)

        for col, (val, fmt) in enumerate(zip(row_data, fmts_row), 1):
            c = ws2.cell(row=rr, column=col, value=val)
            if fmt: c.number_format = fmt
            c.border = _border_thin()
            c.fill   = _fill(bg if usado else "F4F4F4")
            # Fonte: sem_ref fica acinzentada para indicar exclusão visualmente
            font_color = "999999" if not usado else (fc if col == 14 else "000000")
            c.font = _font(size=9, color=font_color, bold=(col == 14 and usado))
            if col in aligns_right: c.alignment = _right()
            # Coluna "Usado no Cálculo": destaque extra
            if col == 15:
                c.font = _font(
                    size=9, bold=True,
                    color=("1A6B3C" if usado else "C0392B")
                )
                c.alignment = _center()
        ws2.row_dimensions[rr].height = 16

    # ── Linha de rodapé de auditoria ────────────────────────────────────────
    rr_rodape = 5 + len(resultado["classified"])
    ws2.merge_cells(f"A{rr_rodape}:{get_column_letter(n_cols)}{rr_rodape}")
    aud = resultado.get("auditoria", {})
    ws2[f"A{rr_rodape}"] = (
        f"AUDITORIA — Total lançamentos: {aud.get('total', len(resultado['classified']))} | "
        f"Computados: {aud.get('computados', n_usado)} | "
        f"Sem referência: {aud.get('sem_ref', 0)} | "
        f"Excluídos pelo operador: {aud.get('excl_operador', 0)} | "
        f"Fora da competência: {aud.get('fora_comp', 0)} | "
        f"Fat. computável: R$ {resultado['fat_comp']:,.2f} | "
        f"Fat. total: R$ {resultado['fat_total']:,.2f} | "
        f"BeefPassion OTE v1.4 · Stoic Capital"
    )
    ws2[f"A{rr_rodape}"].fill = _fill("1F2D3D")
    ws2[f"A{rr_rodape}"].font = _font(size=8, color="DCC697", italic=True)
    ws2[f"A{rr_rodape}"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[rr_rodape].height = 18

    # ── Salvar ─────────────────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(
        suffix=f"_OTE_{vendedor['nome']}_{mes:02d}_{ano}.xlsx",
        delete=False
    )
    wb.save(tmp.name)
    return tmp.name


# ── ROTAS ──────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config", methods=["GET"])
def get_config():
    cfg = load_cfg()
    # Injetar mult_rows: subset exibível da MULT_TABLE (marcos de atingimento)
    marcos = [0.70,0.75,0.80,0.85,0.90,0.95,1.00,1.05,1.10,1.15,1.20,
              1.25,1.30,1.35,1.40,1.50,1.60,1.70,1.80,1.90,2.00,2.50]
    mult_override = cfg.get("mult_table", {})
    mult_rows = {}
    for marco in marcos:
        key = f"{marco:.2f}"
        best_row = MULT_TABLE[0]
        for row in MULT_TABLE:
            if row[0] <= marco:
                best_row = row
            else:
                break
        base = {"c1": best_row[1], "c2": best_row[2], "c3": best_row[3], "c4": best_row[4]}
        mult_rows[key] = mult_override.get(key, base)
    cfg["mult_rows"] = mult_rows
    return jsonify(cfg)


@app.route("/api/config", methods=["POST"])
def post_config():
    data = request.get_json()
    cfg = load_cfg()
    global _MATCH_INDEX
    _MATCH_INDEX = {}  # Invalida cache de matching ao atualizar preços
    if "precos_pj" in data:
        cfg["precos_pj"] = data["precos_pj"]
    if "precos_pf" in data:
        cfg["precos_pf"] = data["precos_pf"]
    if "vendedores" in data:
        cfg["vendedores"] = data["vendedores"]
    if "ote" in data:
        cfg["ote"] = data["ote"]
    if "depara" in data:
        cfg["depara"] = data["depara"]
    if "mult_table" in data:
        cfg["mult_table"] = data["mult_table"]
    if "senha" in data:
        cfg["senha"] = data["senha"]
    save_cfg(cfg)
    return jsonify({"ok": True})


@app.route("/api/extrair", methods=["POST"])
def extrair():
    if "pdf" not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    f = request.files["pdf"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"erro": "Arquivo deve ser PDF."}), 400

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    f.save(tmp.name)
    tmp.close()

    try:
        resultado = extrair_itens(tmp.name)
    except Exception as e:
        return jsonify({"erro": f"Falha na extração: {str(e)}"}), 500
    finally:
        os.unlink(tmp.name)

    return jsonify(resultado)


@app.route("/api/calcular", methods=["POST"])
def calcular_route():
    data = request.get_json()
    itens    = data.get("itens", [])
    canal    = data.get("canal", "PJ")
    nivel    = int(data.get("nivel", 3))
    mes      = int(data.get("mes", 1))
    ano      = int(data.get("ano", 2026))
    nome_vend= data.get("vendedor", "")

    cfg = load_cfg()

    ote_tabela = cfg["ote"].get(canal, [])
    ote_row = next((r for r in ote_tabela if r["n"] == nivel), None)
    if not ote_row:
        return jsonify({"erro": f"Nível {nivel} não encontrado para canal {canal}."}), 400

    vendedor = next((v for v in cfg["vendedores"] if v["nome"] == nome_vend), None)
    if not vendedor:
        vendedor = {"nome": nome_vend, "canal": canal, "cargo": canal}

    # Seleciona tabela de preços pelo canal do vendedor
    precos_canal = cfg.get("precos_pj", {}) if canal == "PJ" else cfg.get("precos_pf", {})

    resultado = calcular_ote(itens, ote_row, precos_canal, canal)
    resultado["ote_row"] = ote_row
    resultado["vendedor"] = vendedor
    resultado["mes"] = mes
    resultado["ano"] = ano
    resultado["nivel"] = nivel

    # Serializar (remover itens completos para não pesar demais — frontend já tem)
    resultado_json = {k: v for k, v in resultado.items() if k != "classified"}
    resultado_json["classified"] = resultado["classified"]
    resultado_json["fx_totais"] = {
        f: {"fat": resultado["fx"][f]["fat"], "n_itens": len(resultado["fx"][f]["itens"])}
        for f in resultado["fx"]
    }

    return jsonify(resultado_json)


@app.route("/api/exportar_xlsx", methods=["POST"])
def exportar_xlsx():
    data = request.get_json()

    cfg      = load_cfg()
    canal    = data.get("canal", "PJ")
    nivel    = int(data.get("nivel", 3))
    mes      = int(data.get("mes", 1))
    ano      = int(data.get("ano", 2026))
    nome_vend= data.get("vendedor", "")
    itens    = data.get("itens", [])

    ote_tabela = cfg["ote"].get(canal, [])
    ote_row = next((r for r in ote_tabela if r["n"] == nivel), None)
    if not ote_row:
        return jsonify({"erro": "Nível OTE não encontrado."}), 400

    vendedor = next((v for v in cfg["vendedores"] if v["nome"] == nome_vend),
                    {"nome": nome_vend, "canal": canal, "cargo": canal})

    # Seleciona tabela de preços pelo canal do vendedor
    precos_canal = cfg.get("precos_pj", {}) if canal == "PJ" else cfg.get("precos_pf", {})

    resultado = calcular_ote(itens, ote_row, precos_canal, canal)
    resultado["ote_row"] = ote_row

    xlsx_path = gerar_xlsx(resultado, vendedor, nivel, mes, ano)

    MESES = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    nome_arquivo = f"OTE_{nome_vend}_{MESES[mes]}{ano}.xlsx"

    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    print("\n" + "="*55)
    print("  BeefPassion · Calculadora OTE  |  Stoic Capital")
    print("="*55)
    print("  Acesse: http://localhost:5000")
    print("  Para encerrar: Ctrl+C")
    print("="*55 + "\n")
    app.run(debug=False, host="127.0.0.1", port=5000)
